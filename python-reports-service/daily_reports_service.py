#!/usr/bin/env python3
"""
DDRC Daily Reports Service
Optimized Python service for generating and sending daily email reports
Handles billions of records efficiently using chunking and parallel processing
"""

import os
import sys
import json
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional, Generator
from pathlib import Path
from dataclasses import dataclass
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed
import mysql.connector
from mysql.connector import pooling
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import smtplib
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('daily_reports.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


@dataclass
class ReportFile:
    """Represents a generated report file"""
    type: str
    value: str
    pdf_path: str
    excel_path: str


@dataclass
class EmailRecipient:
    """Represents an email recipient"""
    id: int
    name: str
    email: str
    recipient_type: str  # 'admin' or 'field_officer'


class DatabaseManager:
    """Manages database connections with connection pooling"""
    
    def __init__(self):
        self.config = {
            'host': os.getenv('DB_HOST', 'localhost'),
            'port': int(os.getenv('DB_PORT', 3306)),
            'user': os.getenv('DB_USER'),
            'password': os.getenv('DB_PASSWORD'),
            'database': os.getenv('DB_NAME'),
            'pool_name': 'ddrc_pool',
            'pool_size': int(os.getenv('DB_POOL_SIZE', 10)),
            'pool_reset_session': True,
            'autocommit': False,
        }
        self.pool = pooling.MySQLConnectionPool(**self.config)
        logger.info(f"Database connection pool created with size {self.config['pool_size']}")
    
    def get_connection(self):
        """Get a connection from the pool"""
        return self.pool.get_connection()
    
    def fetch_all_surveys_chunked(self, chunk_size: int = 50000) -> Generator[List[Dict[str, Any]], None, None]:
        """
        Fetch all surveys in chunks to handle billions of records efficiently
        Uses streaming cursor to avoid loading all data into memory
        """
        conn = self.get_connection()
        cursor = conn.cursor(dictionary=True, buffered=False)
        
        try:
            query = """
                SELECT 
                    s.id,
                    s.aadhaar_id,
                    s.user_id,
                    s.source,
                    s.survey_json,
                    s.created_at,
                    sa.aadhar_no,
                    sa.holder_name,
                    sa.gender,
                    sa.dob,
                    sa.taluka,
                    sa.district,
                    sa.address_text,
                    u.name AS user_name,
                    u.contact_number AS user_phone
                FROM surveys s
                INNER JOIN survey_aadhar sa ON sa.id = s.aadhaar_id
                LEFT JOIN users u ON u.id = s.user_id
                WHERE s.survey_json IS NOT NULL
                AND s.survey_json != ''
                ORDER BY s.created_at DESC
            """
            
            cursor.execute(query)
            
            chunk = []
            count = 0
            for row in cursor:
                # Parse JSON fields
                if row.get('survey_json'):
                    try:
                        row['survey_json'] = json.loads(row['survey_json'])
                    except:
                        row['survey_json'] = {}
                
                chunk.append(row)
                count += 1
                
                if len(chunk) >= chunk_size:
                    logger.info(f"Yielding chunk of {len(chunk)} surveys (total processed: {count})")
                    yield chunk
                    chunk = []
            
            # Yield remaining records
            if chunk:
                logger.info(f"Yielding final chunk of {len(chunk)} surveys (total processed: {count})")
                yield chunk
                
        finally:
            cursor.close()
            conn.close()
    
    def get_admin_users(self) -> List[EmailRecipient]:
        """Get all admin users with email addresses"""
        conn = self.get_connection()
        cursor = conn.cursor(dictionary=True)
        
        try:
            query = """
                SELECT id, name, email 
                FROM users 
                WHERE user_type = 'admin' 
                AND email IS NOT NULL 
                AND email != '' 
                AND email LIKE '%@%'
                AND (status = 'active' OR is_active = 1 OR status IS NULL)
                ORDER BY id
            """
            cursor.execute(query)
            results = cursor.fetchall()
            
            return [
                EmailRecipient(
                    id=row['id'],
                    name=row.get('name', 'Admin'),
                    email=row['email'],
                    recipient_type='admin'
                )
                for row in results
            ]
        finally:
            cursor.close()
            conn.close()
    
    def get_field_officers_with_stats(self) -> List[Dict[str, Any]]:
        """Get all field officers with their survey statistics"""
        conn = self.get_connection()
        cursor = conn.cursor(dictionary=True)
        
        try:
            query = """
                SELECT 
                    u.id,
                    u.name,
                    u.email,
                    COALESCE(u.name, u.contact_number, CONCAT('User ', u.id)) AS officer_name,
                    COUNT(DISTINCT CASE WHEN s.no_of_questions_unanswered = 0 THEN s.id END) AS completed_surveys,
                    COUNT(DISTINCT CASE WHEN s.no_of_questions_unanswered > 0 THEN s.id END) AS pending_surveys,
                    COUNT(DISTINCT s.id) AS total_surveys
                FROM users u
                LEFT JOIN user_types ut ON ut.id = u.user_type_id
                LEFT JOIN surveys s ON s.user_id = u.id
                WHERE (
                    LOWER(COALESCE(NULLIF(u.user_type, ''), ut.user_type, '')) IN ('field_officer', 'field officer', 'officer')
                )
                AND (u.status = 'active' OR u.is_active = 1 OR u.status IS NULL)
                AND u.email IS NOT NULL 
                AND u.email != ''
                GROUP BY u.id, u.name, u.email, u.contact_number
                ORDER BY u.name
            """
            cursor.execute(query)
            return cursor.fetchall()
        finally:
            cursor.close()
            conn.close()
    
    def get_overall_stats(self) -> Dict[str, Any]:
        """Get overall survey statistics"""
        conn = self.get_connection()
        cursor = conn.cursor(dictionary=True)
        
        try:
            query = """
                SELECT 
                    COUNT(DISTINCT s.id) AS total_surveys,
                    COUNT(DISTINCT CASE WHEN s.no_of_questions_unanswered = 0 THEN s.id END) AS completed_surveys,
                    COUNT(DISTINCT CASE WHEN s.no_of_questions_unanswered > 0 THEN s.id END) AS pending_surveys,
                    COUNT(DISTINCT u.id) AS total_field_officers
                FROM surveys s
                LEFT JOIN users u ON s.user_id = u.id
                LEFT JOIN user_types ut ON ut.id = u.user_type_id
                WHERE (
                    LOWER(COALESCE(NULLIF(u.user_type, ''), ut.user_type, '')) IN ('field_officer', 'field officer', 'officer')
                    OR s.source = 'Divyang Self'
                )
            """
            cursor.execute(query)
            result = cursor.fetchone()
            return result or {
                'total_surveys': 0,
                'completed_surveys': 0,
                'pending_surveys': 0,
                'total_field_officers': 0,
            }
        finally:
            cursor.close()
            conn.close()


class ReportGenerator:
    """Generates PDF and Excel reports efficiently"""
    
    def __init__(self, storage_path: str):
        self.storage_path = Path(storage_path)
        self.storage_path.mkdir(parents=True, exist_ok=True)
    
    def generate_report_filename(self, filter_type: str, filter_value: str, date: datetime, extension: str) -> str:
        """Generate standardized report filename"""
        safe_value = "".join(c for c in filter_value if c.isalnum() or c in (' ', '-', '_')).strip()
        safe_value = safe_value.replace(' ', '-')[:50]
        date_str = date.strftime('%Y-%m-%d')
        return f"{filter_type}-{safe_value}-{date_str}.{extension}"
    
    def extract_filter_data(self, survey: Dict[str, Any]) -> Dict[str, str]:
        """Extract filter data from survey JSON"""
        survey_json = survey.get('survey_json', {})
        if isinstance(survey_json, str):
            try:
                survey_json = json.loads(survey_json)
            except:
                survey_json = {}
        
        filters = {
            'source': survey.get('source', 'Divyang Self'),
            'taluka': survey.get('taluka', 'इतर'),
            'district': survey.get('district', 'इतर'),
            'gender': survey.get('gender', 'निर्दिष्ट नाही'),
            'fieldOfficerName': survey.get('user_name'),
            'udid': 'निर्दिष्ट नाही',  # Extract from survey_json if available
        }
        
        # Extract disability from survey_json
        if isinstance(survey_json, dict):
            # Add logic to extract disability type from survey_json
            filters['disability'] = survey_json.get('disability_type', 'निर्दिष्ट नाही')
        
        return filters
    
    def generate_excel_report(self, surveys: List[Dict[str, Any]], title: str, filename: str) -> str:
        """Generate Excel report using openpyxl for better performance"""
        filepath = self.storage_path / filename
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Survey Report"
        
        # Header style
        header_fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = ['ID', 'Aadhar No', 'Name', 'Gender', 'DOB', 'Taluka', 'District', 'Source', 'Field Officer', 'Created At']
        ws.append(headers)
        
        # Apply header style
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Add data rows
        for survey in surveys:
            row = [
                survey.get('id'),
                survey.get('aadhar_no', ''),
                survey.get('holder_name', ''),
                survey.get('gender', ''),
                survey.get('dob', ''),
                survey.get('taluka', ''),
                survey.get('district', ''),
                survey.get('source', ''),
                survey.get('user_name', ''),
                survey.get('created_at', ''),
            ]
            ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
        logger.info(f"Excel report generated: {filepath}")
        return str(filepath)
    
    def generate_pdf_report(self, surveys: List[Dict[str, Any]], title: str, filename: str) -> str:
        """Generate PDF report using ReportLab"""
        filepath = self.storage_path / filename
        
        doc = SimpleDocTemplate(str(filepath), pagesize=A4)
        elements = []
        
        styles = getSampleStyleSheet()
        
        # Title
        title_para = Paragraph(f"<b>{title}</b>", styles['Title'])
        elements.append(title_para)
        elements.append(Spacer(1, 12))
        
        # Summary
        summary_text = f"Total Surveys: {len(surveys)}"
        summary_para = Paragraph(summary_text, styles['Normal'])
        elements.append(summary_para)
        elements.append(Spacer(1, 12))
        
        # Table data
        table_data = [['ID', 'Aadhar', 'Name', 'Taluka', 'District', 'Source']]
        
        # Add survey data (limit to first 1000 for PDF to avoid huge files)
        for survey in surveys[:1000]:
            table_data.append([
                str(survey.get('id', '')),
                str(survey.get('aadhar_no', ''))[:12],
                str(survey.get('holder_name', ''))[:30],
                str(survey.get('taluka', ''))[:20],
                str(survey.get('district', ''))[:20],
                str(survey.get('source', ''))[:20],
            ])
        
        if len(surveys) > 1000:
            table_data.append(['...', f'... and {len(surveys) - 1000} more records', '', '', '', ''])
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E88E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        doc.build(elements)
        logger.info(f"PDF report generated: {filepath}")
        return str(filepath)


class EmailService:
    """Handles email sending with attachments"""
    
    def __init__(self):
        self.smtp_host = os.getenv('SMTP_HOST', 'smtp.gmail.com')
        self.smtp_port = int(os.getenv('SMTP_PORT', 587))
        self.smtp_user = os.getenv('SMTP_USER')
        self.smtp_password = os.getenv('SMTP_PASSWORD')
        self.email_from = os.getenv('EMAIL_FROM', 'DDRC Survey System <noreply@ddrc.org>')
    
    def send_email(
        self,
        recipient: EmailRecipient,
        subject: str,
        html_body: str,
        attachments: Optional[List[str]] = None
    ) -> bool:
        """Send email with optional attachments"""
        try:
            msg = MIMEMultipart()
            msg['From'] = self.email_from
            msg['To'] = recipient.email
            msg['Subject'] = subject
            
            msg.attach(MIMEText(html_body, 'html'))
            
            # Add attachments
            if attachments:
                for filepath in attachments:
                    if os.path.exists(filepath):
                        with open(filepath, 'rb') as f:
                            part = MIMEBase('application', 'octet-stream')
                            part.set_payload(f.read())
                            encoders.encode_base64(part)
                            part.add_header(
                                'Content-Disposition',
                                f'attachment; filename= {os.path.basename(filepath)}'
                            )
                            msg.attach(part)
            
            # Send email
            with smtplib.SMTP(self.smtp_host, self.smtp_port) as server:
                server.starttls()
                server.login(self.smtp_user, self.smtp_password)
                server.send_message(msg)
            
            logger.info(f"Email sent successfully to {recipient.email}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to send email to {recipient.email}: {str(e)}")
            return False


def generate_all_reports(db: DatabaseManager, report_gen: ReportGenerator, report_date: datetime) -> List[ReportFile]:
    """
    Generate all filtered reports efficiently using chunked processing
    Handles billions of records by processing in chunks
    """
    logger.info("Starting report generation...")
    generated_files = []
    
    # Collect all surveys in chunks and group by filters
    source_groups = {}
    taluka_groups = {}
    district_groups = {}
    disability_groups = {}
    gender_groups = {}
    officer_groups = {}
    udid_groups = {}
    
    total_processed = 0
    
    # Process surveys in chunks
    for chunk in db.fetch_all_surveys_chunked(chunk_size=int(os.getenv('CHUNK_SIZE', 50000))):
        total_processed += len(chunk)
        logger.info(f"Processing chunk of {len(chunk)} surveys (total: {total_processed})")
        
        for survey in chunk:
            filters = report_gen.extract_filter_data(survey)
            
            # Group by source
            source = filters.get('source', 'Divyang Self')
            source_groups.setdefault(source, []).append(survey)
            
            # Group by taluka
            taluka = filters.get('taluka', 'इतर')
            taluka_groups.setdefault(taluka, []).append(survey)
            
            # Group by district
            district = filters.get('district', 'इतर')
            district_groups.setdefault(district, []).append(survey)
            
            # Group by disability
            disability = filters.get('disability', 'निर्दिष्ट नाही')
            disability_groups.setdefault(disability, []).append(survey)
            
            # Group by gender
            gender = filters.get('gender', 'निर्दिष्ट नाही')
            gender_groups.setdefault(gender, []).append(survey)
            
            # Group by field officer
            if filters.get('fieldOfficerName'):
                officer_name = filters['fieldOfficerName']
                officer_groups.setdefault(officer_name, []).append(survey)
            
            # Group by UDID
            udid = filters.get('udid', 'निर्दिष्ट नाही')
            udid_groups.setdefault(udid, []).append(survey)
    
    logger.info(f"Finished grouping {total_processed} surveys into filter groups")
    
    # Generate reports for each group (can be parallelized)
    max_workers = int(os.getenv('MAX_WORKERS', 4))
    
    def generate_group_reports(group_type: str, groups: Dict[str, List[Dict]]):
        files = []
        for value, surveys in groups.items():
            try:
                pdf_filename = report_gen.generate_report_filename(group_type, value, report_date, 'pdf')
                excel_filename = report_gen.generate_report_filename(group_type, value, report_date, 'xlsx')
                
                pdf_path = report_gen.generate_pdf_report(surveys, f"{group_type}: {value}", pdf_filename)
                excel_path = report_gen.generate_excel_report(surveys, f"{group_type}: {value}", excel_filename)
                
                files.append(ReportFile(
                    type=group_type.lower().replace(' ', '_'),
                    value=value,
                    pdf_path=pdf_path,
                    excel_path=excel_path
                ))
            except Exception as e:
                logger.error(f"Error generating report for {group_type} - {value}: {str(e)}")
        return files
    
    # Generate reports in parallel
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(generate_group_reports, 'Source', source_groups): 'source',
            executor.submit(generate_group_reports, 'Taluka', taluka_groups): 'taluka',
            executor.submit(generate_group_reports, 'District', district_groups): 'district',
            executor.submit(generate_group_reports, 'Disability', disability_groups): 'disability',
            executor.submit(generate_group_reports, 'Gender', gender_groups): 'gender',
            executor.submit(generate_group_reports, 'Field-Officer', officer_groups): 'field_officer',
            executor.submit(generate_group_reports, 'UDID', udid_groups): 'udid',
        }
        
        for future in as_completed(futures):
            try:
                files = future.result()
                generated_files.extend(files)
                logger.info(f"Generated {len(files)} reports for {futures[future]}")
            except Exception as e:
                logger.error(f"Error in report generation: {str(e)}")
    
    logger.info(f"Total reports generated: {len(generated_files)}")
    return generated_files


def generate_email_body(stats: Dict[str, Any], officers: List[Dict], date_str: str, report_count: int = 0, is_admin: bool = True) -> str:
    """Generate HTML email body"""
    if is_admin:
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                .container {{ max-width: 800px; margin: 0 auto; padding: 20px; }}
                .header {{ background: linear-gradient(135deg, #0D47A1 0%, #1976D2 50%, #42A5F5 100%); color: white; padding: 20px; border-radius: 8px 8px 0 0; }}
                .content {{ background: #f9f9f9; padding: 20px; border: 1px solid #ddd; border-top: none; }}
                .stats-grid {{ display: grid; grid-template-columns: repeat(2, 1fr); gap: 15px; margin-top: 20px; }}
                .stat-card {{ background: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
                .stat-value {{ font-size: 32px; font-weight: bold; color: #1976D2; }}
                .stat-label {{ color: #666; margin-top: 5px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1 style="margin: 0;">DDRC Survey System - Daily Report</h1>
                    <p style="margin: 10px 0 0 0; opacity: 0.9;">{date_str}</p>
                </div>
                <div class="content">
                    <h2 style="color: #1976D2;">Overall Statistics</h2>
                    <div class="stats-grid">
                        <div class="stat-card">
                            <div class="stat-value">{stats.get('total_surveys', 0)}</div>
                            <div class="stat-label">Total Surveys</div>
                        </div>
                        <div class="stat-card">
                            <div class="stat-value">{stats.get('completed_surveys', 0)}</div>
                            <div class="stat-label">Completed Surveys</div>
                        </div>
                        <div class="stat-card">
                            <div class="stat-value">{stats.get('pending_surveys', 0)}</div>
                            <div class="stat-label">Pending Surveys</div>
                        </div>
                        <div class="stat-card">
                            <div class="stat-value">{stats.get('total_field_officers', 0)}</div>
                            <div class="stat-label">Active Field Officers</div>
                        </div>
                    </div>
                    {f'<p style="margin-top: 30px;">This email includes {report_count} filtered report files (PDF and Excel formats).</p>' if report_count > 0 else ''}
                    <p style="margin-top: 30px; color: #666; font-size: 14px;">
                        This is an automated daily report from the DDRC Survey System.
                    </p>
                </div>
            </div>
        </body>
        </html>
        """
    else:
        # Field officer email body
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                .header {{ background: linear-gradient(135deg, #0D47A1 0%, #1976D2 50%, #42A5F5 100%); color: white; padding: 20px; border-radius: 8px 8px 0 0; }}
                .content {{ background: #f9f9f9; padding: 20px; border: 1px solid #ddd; border-top: none; }}
                .stats-grid {{ display: grid; grid-template-columns: repeat(2, 1fr); gap: 15px; margin-top: 20px; }}
                .stat-card {{ background: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); text-align: center; }}
                .stat-value {{ font-size: 36px; font-weight: bold; color: #1976D2; }}
                .stat-label {{ color: #666; margin-top: 8px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1 style="margin: 0;">DDRC Survey System - Daily Report</h1>
                    <p style="margin: 10px 0 0 0; opacity: 0.9;">{date_str}</p>
                </div>
                <div class="content">
                    <h2 style="color: #1976D2;">Your Daily Statistics</h2>
                    <p>This is an automated daily report from the DDRC Survey System.</p>
                </div>
            </div>
        </body>
        </html>
        """


def main():
    """Main function to run the daily reports service"""
    logger.info("=" * 80)
    logger.info("DDRC Daily Reports Service - Starting")
    logger.info("=" * 80)
    
    try:
        # Initialize services
        db = DatabaseManager()
        storage_path = os.getenv('REPORTS_STORAGE_PATH', './storage/reports')
        report_gen = ReportGenerator(storage_path)
        email_service = EmailService()
        
        report_date = datetime.now()
        date_str = report_date.strftime('%A, %d %B %Y')
        
        # Get recipients
        logger.info("Fetching admin users...")
        admins = db.get_admin_users()
        logger.info(f"Found {len(admins)} admin users")
        
        logger.info("Fetching field officers...")
        officers = db.get_field_officers_with_stats()
        logger.info(f"Found {len(officers)} field officers")
        
        logger.info("Fetching overall statistics...")
        stats = db.get_overall_stats()
        logger.info(f"Statistics: {stats}")
        
        # Generate all reports
        logger.info("Generating reports...")
        report_files = generate_all_reports(db, report_gen, report_date)
        logger.info(f"Generated {len(report_files)} report files")
        
        # Prepare attachments for admin emails
        attachments = []
        for report_file in report_files:
            attachments.append(report_file.pdf_path)
            attachments.append(report_file.excel_path)
        
        # Send emails to admins
        logger.info(f"Sending emails to {len(admins)} admins...")
        admin_success = 0
        admin_failed = 0
        
        for admin in admins:
            email_body = generate_email_body(stats, officers, date_str, len(report_files), is_admin=True)
            subject = f"DDRC Survey Daily Report - {date_str}"
            
            if email_service.send_email(admin, subject, email_body, attachments):
                admin_success += 1
            else:
                admin_failed += 1
        
        # Send emails to field officers
        logger.info(f"Sending emails to {len(officers)} field officers...")
        officer_success = 0
        officer_failed = 0
        
        for officer in officers:
            recipient = EmailRecipient(
                id=officer['id'],
                name=officer.get('name', 'Field Officer'),
                email=officer['email'],
                recipient_type='field_officer'
            )
            email_body = generate_email_body(stats, [officer], date_str, 0, is_admin=False)
            subject = f"DDRC Survey Daily Report - {date_str}"
            
            if email_service.send_email(recipient, subject, email_body):
                officer_success += 1
            else:
                officer_failed += 1
        
        logger.info("=" * 80)
        logger.info("Daily Reports Service - Completed")
        logger.info(f"Admins: {admin_success} sent, {admin_failed} failed")
        logger.info(f"Officers: {officer_success} sent, {officer_failed} failed")
        logger.info(f"Total reports generated: {len(report_files)}")
        logger.info("=" * 80)
        
    except Exception as e:
        logger.error(f"Fatal error in daily reports service: {str(e)}", exc_info=True)
        sys.exit(1)


if __name__ == '__main__':
    main()

